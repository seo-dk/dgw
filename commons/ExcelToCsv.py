import csv
import logging
import os
import sys
from functools import wraps
from pathlib import Path
from time import perf_counter

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def timer_decorator(func):
    @wraps(func)
    def wrapper(self, *args, **kwargs):
        local_logger = logging.getLogger(f"[pid: {os.getpid()}]")
        start_time = perf_counter()
        result = func(self, *args, **kwargs)
        end_time = perf_counter()
        local_logger.info(f"{func.__name__} took {end_time - start_time:.2f}s")
        return result

    return wrapper


class ExcelToCsv:

    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path
        self.directory = os.path.dirname(excel_file_path)
        self.filename = os.path.splitext(os.path.basename(excel_file_path))[0]

    def _write_sheet_to_csv(self, sheet_name, sheet_index, workbook, local_logger):
        worksheet = workbook[sheet_name]
        worksheet.reset_dimensions()
        csv_path = Path(self.directory, f"{self.filename}_{sheet_index}.csv")

        with csv_path.open('w', newline='', encoding="utf-8") as csvfile:
            csv_writer = csv.writer(csvfile, delimiter='\036')
            rows_to_write = []

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                rows_to_write.append(row)
                if len(rows_to_write) >= 5000:
                    csv_writer.writerows(rows_to_write)
                    rows_to_write.clear()

            if rows_to_write:
                csv_writer.writerows(rows_to_write)

        local_logger.debug('csv file written to %s', csv_path)
        return str(csv_path)

    # @timer_decorator
    def _process_sheet(self, args):
        local_logger = logging.getLogger(f"[pid: {os.getpid()}]")
        sheet_name, sheet_index = args
        try:
            workbook = load_workbook(self.excel_file_path, read_only=True, data_only=True)
            self._write_sheet_to_csv(sheet_name, sheet_index, workbook, local_logger)
            csv_path = self._write_sheet_to_csv(sheet_name, sheet_index, workbook, local_logger)
            return csv_path
            # return True
        except Exception as e:
            local_logger.exception("sheet_name:%s, sheet_index:%s, exception: %s", sheet_name, sheet_index, e)
            # return False
            return None

    def convert(self):
        file_size = os.path.getsize(self.excel_file_path)
        if file_size < 10:
            logger.warning("File size is less than 10 bytes, the file may not be valid.")
            return []

        logger.info("Starting parallel conversion from Excel to CSV")
        workbook = load_workbook(self.excel_file_path, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames
        sheet_args = [(name, index) for index, name in enumerate(sheet_names, 1)]

        results = [self._process_sheet(sheet_args[0])]

        if len(results) == len(sheet_names) and all(success for success in results):
            logger.info("All sheets processed successfully")
            return results
        else:
            empty_paths = [index for index, value in enumerate(results) if value is None]
            for index in empty_paths:
                logger.info("sheet %s did not converted properly", index)
            results = [value for value in results if value is not None]

        logger.info("xlsx to csv convert finished. excel_file_path: %s", self.excel_file_path)
        return results

if __name__ == "__main__":

    if len(sys.argv) > 1:
        try:
            results = []
            results = ExcelToCsv(sys.argv[1]).convert()
            logger.info("Results : %s", results)
        except Exception as e:
            logger.exception("exception: %s", e)
    else:
        logger.warning("usage: ExcelToCsv.py file.xlsx")
